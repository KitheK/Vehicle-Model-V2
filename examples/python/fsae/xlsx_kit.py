"""Standard .xlsx workbooks for FSAE cars, maps, and driver / 6DOF lap channels.

Three workbook kinds (detected from sheet names, or a ``Format`` sheet):

**Car** (OpenVEHICLE)
  ``Info`` — Category | Description | Value | Unit | Comment
  ``Torque Curve`` — Engine Speed [rpm] | Torque [Nm]
  ``FastestLap`` — same five-column layout for 6DOF fields OpenVEHICLE omits

**Map** (OpenTRACK)
  ``Info`` — Name, Country, City, Type, Configuration, Direction, Mirror
  ``Shape`` — Type | Section Length | Corner Radius  (Straight / Left / Right)

**Driver** (lap + 6DOF channels)
  ``Info`` — Vehicle, Track, Lap time [s], Notes
  ``Channels`` — one row per mesh point (see ``DRIVER_COLUMNS``)
  ``Envelope`` — optional G-G table (ay, ax_max, ax_min)
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from fsae.qss_channels import LapView

FORMAT_ROWS = (
    ("Workbook", "Sheets", "Purpose"),
    ("Car", "Info, Torque Curve, FastestLap", "OpenVEHICLE vehicle + 6DOF FastestLap fields"),
    ("Map", "Info, Shape", "OpenTRACK racing-line geometry"),
    ("Driver", "Info, Channels, Envelope", "Lap, driver inputs, and reconstructed 6DOF channels"),
)

DRIVER_COLUMNS: Sequence[tuple[str, str]] = (
    ("distance_m", "m"),
    ("time_s", "s"),
    ("x_m", "m"),
    ("y_m", "m"),
    ("yaw_rad", "rad"),
    ("kappa_1pm", "1/m"),
    ("speed_mps", "m/s"),
    ("speed_kmh", "km/h"),
    ("LonAcc_g", "g"),
    ("LatAcc_g", "g"),
    ("GSum_g", "g"),
    ("tps", "-"),
    ("bps", "-"),
    ("steer_deg", "deg"),
    ("delta_deg", "deg"),
    ("beta_deg", "deg"),
    ("Fz_FL_N", "N"),
    ("Fz_FR_N", "N"),
    ("Fz_RL_N", "N"),
    ("Fz_RR_N", "N"),
    ("P_FL_W", "W"),
    ("P_FR_W", "W"),
    ("P_RL_W", "W"),
    ("P_RR_W", "W"),
    ("E_FL_J", "J"),
    ("E_FR_J", "J"),
    ("E_RL_J", "J"),
    ("E_RR_J", "J"),
)


def stamp_format(wb: Workbook, kind: str) -> None:
    """Append a ``Format`` sheet (Kind = car | map | driver). Last sheet, not first."""
    if "Format" in wb.sheetnames:
        return
    ws = wb.create_sheet("Format")
    ws.append(["Kind", kind.lower()])
    ws.append([])
    for row in FORMAT_ROWS:
        ws.append(list(row))
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 56


def detect_kind(path: str | Path) -> str:
    """Return ``car``, ``map``, or ``driver`` from workbook sheets."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        names = set(wb.sheetnames)
        if "Format" in names:
            for row in wb["Format"].iter_rows(max_row=4, values_only=True):
                if row and str(row[0]).strip().lower() == "kind" and row[1]:
                    kind = str(row[1]).strip().lower()
                    if kind in {"car", "map", "driver"}:
                        return kind
        if "Channels" in names:
            return "driver"
        if "Shape" in names:
            return "map"
        if "Torque Curve" in names or "FastestLap" in names:
            return "car"
        if "Info" in names:
            return "car"
    finally:
        wb.close()
    raise ValueError(f"{path} is not a Car, Map, or Driver workbook")


def _channel_row(view: LapView, i: int) -> List[float]:
    gsum = (view.ax[i] ** 2 + view.ay[i] ** 2) ** 0.5
    return [
        view.s[i],
        view.time[i],
        view.x[i],
        view.y[i],
        view.yaw[i],
        view.kappa[i],
        view.v[i],
        view.v[i] * 3.6,
        view.ax[i],
        view.ay[i],
        gsum,
        view.tps[i],
        view.bps[i],
        view.steer[i],
        view.delta[i],
        view.beta[i],
        view.fz_fl[i],
        view.fz_fr[i],
        view.fz_rl[i],
        view.fz_rr[i],
        view.power_fl[i],
        view.power_fr[i],
        view.power_rl[i],
        view.power_rr[i],
        view.energy_fl[i],
        view.energy_fr[i],
        view.energy_rl[i],
        view.energy_rr[i],
    ]


def write_driver_xlsx(view: LapView, path: str | Path) -> Path:
    """Write a Driver workbook (Info + Channels + Envelope) for MATLAB / Excel."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    info = wb.active
    info.title = "Info"
    for row in (
        ("Workbook", "Driver"),
        ("Vehicle", view.vehicle_name),
        ("Track", view.track_name),
        ("Lap time [s]", round(view.lap_time, 4)),
        ("Samples", len(view.s)),
        ("Notes", view.notes),
    ):
        info.append(list(row))
    info.column_dimensions["A"].width = 16
    info.column_dimensions["B"].width = 80

    ch = wb.create_sheet("Channels")
    ch.append([name for name, _ in DRIVER_COLUMNS])
    ch.append([unit for _, unit in DRIVER_COLUMNS])
    for i in range(len(view.s)):
        ch.append(_channel_row(view, i))
    last_col = get_column_letter(len(DRIVER_COLUMNS))
    ch.freeze_panes = "A3"
    ch.auto_filter.ref = f"A1:{last_col}{ch.max_row}"

    env = wb.create_sheet("Envelope")
    env.append(["ay_g", "ax_max_g", "ax_min_g", "speed_mps"])
    n = min(len(view.env_ay), len(view.env_ax_max), len(view.env_ax_min))
    for i in range(n):
        env.append([view.env_ay[i], view.env_ax_max[i], view.env_ax_min[i], view.env_speed])

    stamp_format(wb, "driver")
    wb.save(path)
    return path


def read_driver_xlsx(path: str | Path) -> Dict[str, Any]:
    """Read a Driver workbook into column arrays plus Info metadata."""
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        info: Dict[str, Any] = {}
        if "Info" in wb.sheetnames:
            for row in wb["Info"].iter_rows(values_only=True):
                if row and row[0]:
                    info[str(row[0]).strip()] = row[1]
        if "Channels" not in wb.sheetnames:
            raise ValueError(f"{path} has no Channels sheet")
        rows = list(wb["Channels"].iter_rows(values_only=True))
        if len(rows) < 3:
            raise ValueError(f"{path} Channels sheet has no samples")
        headers = [str(h).strip() for h in rows[0]]
        data: Dict[str, List[float]] = {h: [] for h in headers}
        for row in rows[2:]:
            if row is None or row[0] is None:
                continue
            for h, val in zip(headers, row):
                data[h].append(float(val or 0.0))
        return {"info": info, "channels": data}
    finally:
        wb.close()


def convert_xlsx(src: str | Path, dest: Optional[str | Path] = None) -> Path:
    """Convert a Car or Map workbook to XML. Driver workbooks are validated only."""
    from fsae.openvehicle_xlsx import write_xml
    from fsae.opentrack_xlsx import write_discrete_xml

    src = Path(src)
    kind = detect_kind(src)
    if kind == "car":
        out = Path(dest) if dest else src.with_suffix(".xml")
        return write_xml(src, out)
    if kind == "map":
        out = Path(dest) if dest else src.with_suffix(".xml")
        return write_discrete_xml(src, out)
    read_driver_xlsx(src)
    return src


def list_car_fields(path: str | Path) -> List[Dict[str, Any]]:
    """Info + FastestLap rows for the Studio editor (skip header)."""
    wb = load_workbook(path, data_only=True)
    try:
        fields: List[Dict[str, Any]] = []
        for sheet in ("Info", "FastestLap"):
            if sheet not in wb.sheetnames:
                continue
            last_cat = ""
            for i, row in enumerate(wb[sheet].iter_rows(min_row=2, values_only=True), start=2):
                if not row or len(row) < 3 or row[1] is None:
                    continue
                cat = str(row[0]).strip() if row[0] else last_cat
                if row[0]:
                    last_cat = cat
                fields.append(
                    {
                        "sheet": sheet,
                        "row": i,
                        "category": cat,
                        "description": str(row[1]).strip(),
                        "value": row[2],
                        "unit": "" if len(row) < 4 or row[3] is None else str(row[3]),
                        "comment": "" if len(row) < 5 or row[4] is None else str(row[4]),
                    }
                )
        return fields
    finally:
        wb.close()


def _cell_value(raw: Any) -> Any:
    if raw is None or raw == "":
        return raw
    if isinstance(raw, bool) or isinstance(raw, (int, float)):
        return raw
    text = str(raw).strip()
    try:
        number = float(text)
    except ValueError:
        return raw
    if number.is_integer() and "." not in text and "e" not in text.lower():
        return int(number)
    return number


def patch_car_xlsx(src: str | Path, dest: str | Path, overrides: Sequence[Dict[str, Any]]) -> Path:
    """Copy a Car workbook and set Info/FastestLap Value cells by Description."""
    src, dest = Path(src), Path(dest)
    dest.parent.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(src)
    lookup = {}
    for item in overrides:
        sheet = str(item.get("sheet") or "Info")
        desc = str(item.get("description") or "").strip().lower()
        if desc:
            lookup[(sheet, desc)] = _cell_value(item.get("value"))
    for sheet in ("Info", "FastestLap"):
        if sheet not in wb.sheetnames:
            continue
        for row in wb[sheet].iter_rows(min_row=2):
            if row[1].value is None:
                continue
            key = (sheet, str(row[1].value).strip().lower())
            if key in lookup:
                row[2].value = lookup[key]
    stamp_format(wb, "car")
    wb.save(dest)
    wb.close()
    return dest


def gg_from_driver_xlsx(path: str | Path):
    """Build a QSS G-G table from a Driver Envelope sheet, or None."""
    from fsae.qss_lap import GGTable

    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        if "Envelope" not in wb.sheetnames:
            return None
        rows = list(wb["Envelope"].iter_rows(values_only=True))
        ay: List[float] = []
        ax_max: List[float] = []
        ax_min: List[float] = []
        speed = 15.0
        for row in rows[1:]:
            if not row or row[0] is None:
                continue
            ay.append(float(row[0]))
            ax_max.append(float(row[1] or 0.0))
            ax_min.append(float(row[2] or 0.0))
            if len(row) > 3 and row[3] not in (None, ""):
                speed = float(row[3])
        if len(ay) < 2:
            return None
        return GGTable(speeds=[speed], ay=[ay], ax_max=[ax_max], ax_min=[ax_min])
    finally:
        wb.close()


def preview_workbook(path: str | Path) -> Dict[str, Any]:
    """Parse a Car, Map, or Driver .xlsx for the Studio Process button."""
    path = Path(path)
    kind = detect_kind(path)
    if kind == "car":
        fields = list_car_fields(path)
        name = next((f["value"] for f in fields if f["description"] == "Name"), path.stem)
        return {"kind": "car", "name": name, "fields": fields}
    if kind == "map":
        from fsae.opentrack_xlsx import read_opentrack_info, read_opentrack_shape

        info = read_opentrack_info(path)
        shape = [{"type": k, "length": L, "radius": r} for k, L, r in read_opentrack_shape(path)]
        return {
            "kind": "map",
            "name": info.name,
            "info": {
                "name": info.name,
                "country": info.country,
                "city": info.city,
                "type": info.kind,
                "configuration": info.configuration,
                "direction": info.direction,
                "mirror": info.mirror,
            },
            "shape": shape,
            "length_m": sum(s["length"] for s in shape),
        }
    data = read_driver_xlsx(path)
    return {
        "kind": "driver",
        "info": data["info"],
        "samples": len(data["channels"].get("distance_m") or []),
    }
