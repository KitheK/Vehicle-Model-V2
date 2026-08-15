"""Read an OpenVEHICLE-format workbook and emit fsae-6dof XML.

Workbook layout (same as OpenLAP / OpenVEHICLE tmp.xlsx):

- Sheet ``Info``: header row
  ``Category | Description | Value | Unit | Comment``
  Data starts at row 2. MATLAB OpenVEHICLE.m reads columns B–C
  (Description, Value) from row 2.
- Sheet ``Torque Curve``: header
  ``Engine Speed [rpm] | Torque [Nm]``
  Data starts at row 2.

Optional sheet ``FastestLap`` uses the same five-column Info layout for
6DOF fields OpenVEHICLE does not have (tracks, wheel rates, thermal,
kinematics, aero maps). Missing FastestLap cells fall back to the
UBCO 2026 defaults used by the fsae-6dof model.
"""

from __future__ import annotations

import math
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

INFO_HEADER = ("Category", "Description", "Value", "Unit", "Comment")
TORQUE_HEADER = ("Engine Speed [rpm]", "Torque [Nm]")

# OpenVEHICLE Info: Description (col B) → internal key.
INFO_KEYS = {
    "name": "name",
    "type": "type",
    "mass": "mass",
    "mass (kg)": "mass",
    "total mass": "mass",
    "front mass distribution": "df_percent",
    "wheelbase": "wheelbase_mm",
    "wheelbase (m)": "wheelbase_m",
    "cl": "cl_openvehicle",
    "cd": "cd_openvehicle",
    "lift coefficient": "cl_openvehicle",
    "drag coefficient": "cd_openvehicle",
    "steering rack ratio": "steering_rack_ratio",
    "lift coefficient cl": "cl_openvehicle",
    "drag coefficient cd": "cd_openvehicle",
    "cl scale multiplier": "factor_cl",
    "cd scale multiplier": "factor_cd",
    "front aero distribution": "da_percent",
    "frontal area": "area",
    "air density": "rho",
    "disc outer diameter": "br_disc_d_mm",
    "pad height": "br_pad_h_mm",
    "pad friction coefficient": "br_pad_mu",
    "caliper number of pistons": "br_nop",
    "caliper piston diameter": "br_pist_d_mm",
    "master cylinder piston diameter": "br_mast_d_mm",
    "pedal ratio": "br_ped_r",
    "grip factor multiplier": "factor_grip",
    "tyre radius": "tyre_radius_mm",
    "tire radius": "tyre_radius_mm",
    "rolling resistance": "rolling_resistance",
    "longitudinal friction coefficient": "mu_x",
    "longitudinal friction load rating": "mu_x_M",
    "longitudinal friction sensitivity": "sens_x",
    "lateral friction coefficient": "mu_y",
    "lateral friction load rating": "mu_y_M",
    "lateral friction sensitivity": "sens_y",
    "front cornering stiffness": "CF",
    "rear cornering stiffness": "CR",
    "power factor multiplier": "factor_power",
    "thermal efficiency": "n_thermal",
    "fuel lower heating value": "fuel_LHV",
    "drive type": "drive",
    "gear shift time": "shift_time",
    "primary gear efficiency": "n_primary",
    "final gear efficiency": "n_final",
    "gearbox efficiency": "n_gearbox",
    "primary gear reduction": "ratio_primary",
    "final gear reduction": "ratio_final",
    "1st gear ratio": "ratio_gear_1",
}

FASTESTLAP_KEYS = {
    "front track": "front_track_mm",
    "rear track": "rear_track_mm",
    "cg height": "cg_height_mm",
    "front axle z": "front_axle_z_mm",
    "rear axle z": "rear_axle_z_mm",
    "front wheel rate": "front_wheel_rate",
    "front antiroll": "front_antiroll",
    "front damper": "front_damper",
    "rear wheel rate": "rear_wheel_rate",
    "rear antiroll": "rear_antiroll",
    "rear damper": "rear_damper",
    "axle inertia": "axle_inertia",
    "peak motor torque": "peak_torque",
    "maximum power": "maximum_power_kw",
    "maximum power (kw)": "maximum_power_kw",
    "gear ratio": "gear_ratio",
    "regen coefficient": "regen_coefficient",
    "differential stiffness": "differential_stiffness",
    "max brake torque": "max_brake_torque",
    "brake bias": "brake_bias",
    "camber static": "camber_static_deg",
    "camber gain roll": "camber_gain_roll",
    "toe static": "toe_static_deg",
    "toe gain roll": "toe_gain_roll",
    "dcl_dz": "dCl_dz",
    "dcl_dmu": "dCl_dmu",
    "dcd_dz": "dCd_dz",
    "dcd_dmu": "dCd_dmu",
    "tire thermal capacity": "thermal_capacity",
    "tire thermal cooling": "thermal_cooling",
    "tire t ambient": "t_ambient",
    "tire t optimal": "t_optimal",
    "tire grip sensitivity": "grip_sensitivity",
    "ixx": "Ixx",
    "iyy": "Iyy",
    "izz": "Izz",
    "ixz": "Ixz",
    "tire radial stiffness": "tire_radial_stiffness",
    "tire radial damping": "tire_radial_damping",
    "nominal vertical load": "nominal_vertical_load",
}

DEFAULTS = {
    "name": "UBCO 2026 EV",
    "type": "FSAE",
    "mass": 277.2,
    "df_percent": 51.27,
    "wheelbase_mm": 1620.0,
    "cl_openvehicle": -3.77,
    "cd_openvehicle": -1.4,
    "factor_cl": 1.0,
    "factor_cd": 1.0,
    "da_percent": 48.0,
    "area": 1.14,
    "rho": 1.2,
    "tyre_radius_mm": 203.0,
    "mu_x": 0.9,
    "mu_y": 1.5,
    "drive": "RWD",
    "ratio_primary": 1.0,
    "ratio_final": 4.8,
    "ratio_gear_1": 1.0,
    "front_track_mm": 1225.0,
    "rear_track_mm": 1220.0,
    "cg_height_mm": 250.0,
    "front_axle_z_mm": 47.0,
    "rear_axle_z_mm": 47.0,
    "front_wheel_rate": 130000.0,
    "front_antiroll": 11260.0,
    "front_damper": 1500.0,
    "rear_wheel_rate": 143900.0,
    "rear_antiroll": 11160.0,
    "rear_damper": 1400.0,
    "axle_inertia": 0.5,
    "regen_coefficient": 0.3,
    "differential_stiffness": 80.0,
    "max_brake_torque": 1200.0,
    "brake_bias": 0.53,
    "camber_static_deg": -1.0,
    "camber_gain_roll": 0.3,
    "toe_static_deg": 0.0,
    "toe_gain_roll": 0.0,
    "dCl_dz": -8.0,
    "dCl_dmu": 2.0,
    "dCd_dz": -2.0,
    "dCd_dmu": 0.5,
    "thermal_capacity": 900.0,
    "thermal_cooling": 15.0,
    "t_ambient": 298.15,
    "t_optimal": 353.15,
    "grip_sensitivity": 0.3,
    "Ixx": 30.0,
    "Iyy": 90.0,
    "Izz": 110.0,
    "Ixz": 5.0,
    "tire_radial_stiffness": 240000.0,
    "tire_radial_damping": 1000.0,
    "nominal_vertical_load": 680.0,
    "smooth_throttle_coeff": 0.01,
}


def _norm(text: Any) -> str:
    if text is None:
        return ""
    return re.sub(r"\s+", " ", str(text)).strip().lower()


_MM_KEYS = {
    "wheelbase_mm",
    "front_track_mm",
    "rear_track_mm",
    "cg_height_mm",
    "front_axle_z_mm",
    "rear_axle_z_mm",
    "tyre_radius_mm",
    "br_disc_d_mm",
    "br_pad_h_mm",
    "br_pist_d_mm",
    "br_mast_d_mm",
}


def _lookup_key(description: Any, keymap: Dict[str, str]) -> Optional[str]:
    n = _norm(description)
    if not n:
        return None
    if n in keymap:
        return keymap[n]
    stripped = re.sub(r"\s*\([^)]*\)", "", n)
    stripped = re.sub(r"\s+", " ", stripped).strip()
    return keymap.get(stripped)


def _apply_unit(key: str, value: float, unit: Any) -> float:
    u = _norm(unit)
    if not u:
        return value
    if key in _MM_KEYS and u in {"m", "meter", "meters", "metre", "metres"}:
        return value * 1000.0
    if key == "maximum_power_kw" and u in {"w", "watt", "watts"}:
        return value / 1000.0
    if key == "mass" and u in {"g", "gram", "grams"}:
        return value / 1000.0
    return value


def _as_float(value: Any, default: Optional[float] = None) -> Optional[float]:
    if value is None or value == "":
        return default
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return default


def _read_description_sheet(ws: Worksheet, keymap: Dict[str, str]) -> Dict[str, Any]:
    out: Dict[str, Any] = {}
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 1:
            continue
        if row is None or len(row) < 3:
            continue
        description, value = row[1], row[2]
        unit = row[3] if len(row) > 3 else None
        key = _lookup_key(description, keymap)
        if key is None or value is None or value == "":
            continue
        if key in ("name", "type", "drive"):
            out[key] = str(value).strip()
        else:
            parsed = _as_float(value)
            if parsed is not None:
                out[key] = _apply_unit(key, parsed, unit)
    return out


def _read_torque_curve(ws: Worksheet) -> List[Tuple[float, float]]:
    points: List[Tuple[float, float]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 1 or row is None or len(row) < 2:
            continue
        rpm, torque = _as_float(row[0]), _as_float(row[1])
        if rpm is None or torque is None:
            continue
        points.append((rpm, torque))
    return points


def read_openvehicle_xlsx(path: str | Path) -> Dict[str, Any]:
    """Parse an OpenVEHICLE workbook into a dict of physical parameters."""
    path = Path(path)
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        if "Info" not in wb.sheetnames:
            raise ValueError(f"{path} has no 'Info' sheet (OpenVEHICLE format).")
        params = dict(DEFAULTS)
        params.update(_read_description_sheet(wb["Info"], INFO_KEYS))
        if "FastestLap" in wb.sheetnames:
            params.update(_read_description_sheet(wb["FastestLap"], FASTESTLAP_KEYS))
        torque = []
        if "Torque Curve" in wb.sheetnames:
            torque = _read_torque_curve(wb["Torque Curve"])
        params["torque_curve"] = torque
    finally:
        wb.close()
    return params


def derived_vehicle(params: Dict[str, Any]) -> Dict[str, Any]:
    """Map OpenVEHICLE numbers onto fsae-6dof XML fields (SI)."""
    p = dict(DEFAULTS)
    p.update(params)

    mass = float(p["mass"])
    df = float(p["df_percent"]) / 100.0
    if p.get("wheelbase_m") is not None:
        L = float(p["wheelbase_m"])
    else:
        L = float(p["wheelbase_mm"]) / 1000.0
    da = float(p["da_percent"]) / 100.0
    h_cg = float(p["cg_height_mm"]) / 1000.0

    x_front = (1.0 - df) * L
    x_rear = -df * L
    z_front = float(p["front_axle_z_mm"]) / 1000.0
    z_rear = float(p["rear_axle_z_mm"]) / 1000.0
    x_pc = x_rear + da * L

    # OpenVEHICLE: CL>0 lift, CD should be negative. fastest-lap uses +Cl/+Cd.
    cl = -float(p["cl_openvehicle"]) * float(p["factor_cl"])
    cd = -float(p["cd_openvehicle"]) * float(p["factor_cd"])

    gear = p.get("gear_ratio")
    if gear is None:
        gear = float(p["ratio_primary"]) * float(p.get("ratio_gear_1") or 1.0) * float(p["ratio_final"])

    peak_torque = p.get("peak_torque")
    max_power_kw = p.get("maximum_power_kw")
    curve = p.get("torque_curve") or []
    if curve:
        torques = [t for _, t in curve]
        if peak_torque is None:
            peak_torque = max(torques)
        if max_power_kw is None:
            powers_w = [t * rpm * 2.0 * math.pi / 60.0 for rpm, t in curve]
            max_power_kw = max(powers_w) / 1000.0
    if peak_torque is None:
        peak_torque = 240.0
    if max_power_kw is None:
        max_power_kw = 80.0

    radius = float(p["tyre_radius_mm"]) / 1000.0
    camber_static = math.radians(float(p["camber_static_deg"]))
    toe_static = math.radians(float(p["toe_static_deg"]))

    return {
        **p,
        "wheelbase_m": L,
        "df": df,
        "da": da,
        "x_front": x_front,
        "x_rear": x_rear,
        "z_front": z_front,
        "z_rear": z_rear,
        "x_pc": x_pc,
        "h_cg": h_cg,
        "cl": cl,
        "cd": cd,
        "gear_ratio": float(gear),
        "peak_torque": float(peak_torque),
        "maximum_power_kw": float(max_power_kw),
        "tyre_radius_m": radius,
        "camber_static_rad": camber_static,
        "toe_static_rad": toe_static,
        "front_track_m": float(p["front_track_mm"]) / 1000.0,
        "rear_track_m": float(p["rear_track_mm"]) / 1000.0,
    }


def _fmt(value: float, digits: int = 6) -> str:
    text = f"{value:.{digits}f}".rstrip("0").rstrip(".")
    if text in ("-0", "-0.0"):
        return "0"
    return text


def vehicle_to_xml(params: Dict[str, Any]) -> str:
    v = derived_vehicle(params)
    camber = _fmt(v["camber_static_rad"], 5)
    toe = _fmt(v["toe_static_rad"], 5)
    return f"""<!--
  {v['name']} ({v['type']}), generated from an OpenVEHICLE-format .xlsx.

  Geometry, mass, aero, and motor numbers come from the workbook Info /
  Torque Curve / FastestLap sheets. Pacejka shape coefficients that are
  not in OpenVEHICLE stay as placeholders until Hoosier TTC data is fitted.
-->
<vehicle type="fsae-6dof">
    <front-axle model="axle-car">
        <track units="m">{_fmt(v['front_track_m'], 4)}</track>
        <stiffness>
            <wheel-rate>{_fmt(v['front_wheel_rate'], 1)}</wheel-rate>
            <antiroll>{_fmt(v['front_antiroll'], 1)}</antiroll>
            <damper>{_fmt(v['front_damper'], 1)}</damper>
        </stiffness>
        <inertia>{_fmt(v['axle_inertia'], 2)}</inertia>
        <smooth_throttle_coeff>{_fmt(v['smooth_throttle_coeff'], 4)}</smooth_throttle_coeff>
        <beta-steering>
            <left>0.0</left>
            <right>0.0</right>
        </beta-steering>
        <kinematics>
            <camber_static>{camber}</camber_static>
            <camber_gain_roll>{_fmt(v['camber_gain_roll'], 4)}</camber_gain_roll>
            <toe_static>{toe}</toe_static>
            <toe_gain_roll>{_fmt(v['toe_gain_roll'], 4)}</toe_gain_roll>
        </kinematics>
        <brakes>
            <max_torque units="N.m">{_fmt(v['max_brake_torque'], 1)}</max_torque>
        </brakes>
    </front-axle>

    <rear-axle model="axle-car">
        <track units="m">{_fmt(v['rear_track_m'], 4)}</track>
        <stiffness>
            <wheel-rate>{_fmt(v['rear_wheel_rate'], 1)}</wheel-rate>
            <antiroll>{_fmt(v['rear_antiroll'], 1)}</antiroll>
            <damper>{_fmt(v['rear_damper'], 1)}</damper>
        </stiffness>
        <inertia>{_fmt(v['axle_inertia'], 2)}</inertia>
        <smooth_throttle_coeff>{_fmt(v['smooth_throttle_coeff'], 4)}</smooth_throttle_coeff>
        <differential_stiffness units="N.m.s/rad">{_fmt(v['differential_stiffness'], 1)}</differential_stiffness>
        <brakes>
            <max_torque units="N.m">{_fmt(v['max_brake_torque'], 1)}</max_torque>
        </brakes>
        <engine>
            <maximum-power units="kW">{_fmt(v['maximum_power_kw'], 2)}</maximum-power>
            <peak-torque units="N.m">{_fmt(v['peak_torque'], 2)}</peak-torque>
            <gear-ratio>{_fmt(v['gear_ratio'], 4)}</gear-ratio>
        </engine>
        <regen_coefficient>{_fmt(v['regen_coefficient'], 4)}</regen_coefficient>
        <kinematics>
            <camber_static>{camber}</camber_static>
            <camber_gain_roll>{_fmt(v['camber_gain_roll'], 4)}</camber_gain_roll>
            <toe_static>{toe}</toe_static>
            <toe_gain_roll>{_fmt(v['toe_gain_roll'], 4)}</toe_gain_roll>
        </kinematics>
    </rear-axle>

    <chassis model="chassis-car">
        <com>0.0 0.0 -{_fmt(v['h_cg'], 4)}</com>
        <front_axle>{_fmt(v['x_front'], 3)} 0.0 {_fmt(v['z_front'], 3)}</front_axle>
        <rear_axle>{_fmt(v['x_rear'], 3)} 0.0 {_fmt(v['z_rear'], 3)}</rear_axle>
        <mass>{_fmt(v['mass'], 4)}</mass>
        <inertia>
            <Ixx> {_fmt(v['Ixx'], 2)} </Ixx>  <Ixy> 0.0  </Ixy> <Ixz> {_fmt(v['Ixz'], 2)}  </Ixz>
            <Iyx> 0.0  </Iyx>  <Iyy> {_fmt(v['Iyy'], 2)} </Iyy> <Iyz> 0.0  </Iyz>
            <Izx> {_fmt(v['Ixz'], 2)}  </Izx>  <Izy> 0.0  </Izy> <Izz> {_fmt(v['Izz'], 2)} </Izz>
        </inertia>
        <aerodynamics>
            <rho>{_fmt(v['rho'], 4)}</rho>
            <cd>{_fmt(v['cd'], 4)}</cd>
            <cl>{_fmt(v['cl'], 4)}</cl>
            <area units="m2">{_fmt(v['area'], 4)}</area>
        </aerodynamics>
        <aero-maps>
            <dCl_dz>{_fmt(v['dCl_dz'], 2)}</dCl_dz>
            <dCl_dmu>{_fmt(v['dCl_dmu'], 2)}</dCl_dmu>
            <dCd_dz>{_fmt(v['dCd_dz'], 2)}</dCd_dz>
            <dCd_dmu>{_fmt(v['dCd_dmu'], 2)}</dCd_dmu>
        </aero-maps>
        <tire-thermal>
            <capacity>{_fmt(v['thermal_capacity'], 2)}</capacity>
            <cooling>{_fmt(v['thermal_cooling'], 2)}</cooling>
            <t-ambient>{_fmt(v['t_ambient'], 2)}</t-ambient>
            <t-optimal>{_fmt(v['t_optimal'], 2)}</t-optimal>
            <grip-sensitivity>{_fmt(v['grip_sensitivity'], 4)}</grip-sensitivity>
        </tire-thermal>
        <pressure_center>{_fmt(v['x_pc'], 3)} 0.0 -{_fmt(v['h_cg'], 4)}</pressure_center>
        <brake_bias>{_fmt(v['brake_bias'], 4)}</brake_bias>
    </chassis>

    <rear-tire model="tire-pacejka" type="normal">
{_tire_block(v)}
    </rear-tire>

    <front-tire model="tire-pacejka" type="normal">
{_tire_block(v)}
    </front-tire>
</vehicle>
"""


def _tire_block(v: Dict[str, Any]) -> str:
    return f"""        <radius>{_fmt(v['tyre_radius_m'], 4)}</radius>
        <radial-stiffness>{_fmt(v['tire_radial_stiffness'], 1)}</radial-stiffness>
        <radial-damping>{_fmt(v['tire_radial_damping'], 1)}</radial-damping>
        <nominal-vertical-load>{_fmt(v['nominal_vertical_load'], 1)}</nominal-vertical-load>
        <lambdaFz0>1.6</lambdaFz0>
        <Fz-max-ref2>1.0</Fz-max-ref2>
        <longitudinal>
            <pure>
                <pCx1>2.3</pCx1>
                <pDx1>{_fmt(v['mu_x'], 4)}</pDx1>
                <pEx1>0.95</pEx1>
                <pKx1>20.0</pKx1>
                <pKx2>1.0</pKx2>
                <pKx3>-0.5</pKx3>
            </pure>
            <combined>
                <rBx1>14.0</rBx1>
                <rCx1>1.0</rCx1>
            </combined>
        </longitudinal>
        <lateral>
            <pure>
                <pCy1>2.3</pCy1>
                <pDy1>{_fmt(v['mu_y'], 4)}</pDy1>
                <pEy1>0.9</pEy1>
                <pKy1>37.6</pKy1>
                <pKy2>1.6</pKy2>
                <pKy4>2.0</pKy4>
            </pure>
            <combined>
                <rBy1>12.0</rBy1>
                <rCy1>0.6</rCy1>
            </combined>
        </lateral>"""


def xlsx_to_xml(xlsx_path: str | Path) -> str:
    return vehicle_to_xml(read_openvehicle_xlsx(xlsx_path))


def write_xml(xlsx_path: str | Path, xml_path: str | Path) -> Path:
    xml_path = Path(xml_path)
    xml_path.parent.mkdir(parents=True, exist_ok=True)
    xml_path.write_text(xlsx_to_xml(xlsx_path), encoding="utf-8")
    return xml_path


def _style_header(ws: Worksheet, n_cols: int) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(bold=True, color="FFFFFF")
    thin = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )
    for col in range(1, n_cols + 1):
        cell = ws.cell(1, col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin


def _write_info_sheet(ws: Worksheet, rows: Sequence[Tuple[Any, ...]]) -> None:
    ws.append(list(INFO_HEADER))
    _style_header(ws, 5)
    alt = PatternFill("solid", fgColor="D6EAF8")
    for i, row in enumerate(rows):
        ws.append(list(row))
        if i % 2 == 1:
            for col in range(1, 6):
                ws.cell(i + 2, col).fill = alt
    widths = [18, 36, 14, 10, 56]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{len(rows) + 1}"


def write_openvehicle_workbook(
    path: str | Path,
    info_rows: Sequence[Tuple[Any, ...]],
    torque_rows: Sequence[Tuple[float, float]],
    fastestlap_rows: Sequence[Tuple[Any, ...]],
) -> Path:
    """Write an OpenVEHICLE-format workbook (Info + Torque Curve + FastestLap)."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    info = wb.active
    info.title = "Info"
    _write_info_sheet(info, info_rows)

    torque = wb.create_sheet("Torque Curve")
    torque.append(list(TORQUE_HEADER))
    _style_header(torque, 2)
    for rpm, tq in torque_rows:
        torque.append([rpm, tq])
    torque.column_dimensions["A"].width = 22
    torque.column_dimensions["B"].width = 16
    torque.freeze_panes = "A2"

    extra = wb.create_sheet("FastestLap")
    _write_info_sheet(extra, fastestlap_rows)

    wb.save(path)
    return path


def ubco_2026_info_rows() -> List[Tuple[Any, ...]]:
    return [
        ("General", "Name", "UBCO 2026 EV", "-", "UBCO Motorsports Formula SAE electric"),
        (None, "Type", "FSAE", "-", None),
        ("Inertia", "Total Mass", 277.2, "kg", "Car 209.2 kg + driver 68 kg"),
        (None, "Front Mass Distribution", 51.27, "%", None),
        ("Dimensions", "Wheelbase", 1620, "mm", None),
        ("Steering", "Steering Rack Ratio", 10, "-", "[Steering Wheel Angle]/[Wheel Angle]"),
        ("Aerodynamics", "Lift Coefficient CL", -3.77, "-", "Positive = Lift / Negative = Downforce"),
        (None, "Drag Coefficient CD", -1.4, "-", "Should be negative"),
        (None, "CL Scale Multiplier", 1, "-", None),
        (None, "CD Scale Multiplier", 1, "-", None),
        (None, "Front Aero Distribution", 48, "%", None),
        (None, "Frontal Area", 1.14, "m2", None),
        (None, "Air Density", 1.2, "kg/m3", "Recommended value:  1.225"),
        ("Brakes", "Disc Outer Diameter", 250, "mm", "Assumed to be the same on all corners."),
        (None, "Pad Height", 40, "mm", "Assumed to be the same on all corners."),
        (None, "Pad Friction Coefficient", 0.45, "-", "Assumed to be the same on all pads."),
        (None, "Caliper Number of Pistons", 4, "-", None),
        (None, "Caliper Piston Diameter", 32, "mm", None),
        (None, "Master Cylinder Piston Diameter", 19, "mm", None),
        (None, "Pedal Ratio", 4, "-", None),
        ("Tyres", "Grip Factor Multiplier", 1, "-", None),
        (None, "Tyre Radius", 203, "mm", "Hoosier 18x7.5-10 class tyre"),
        (None, "Rolling Resistance", -0.001, "-", "Needs to be negative"),
        (None, "Longitudinal Friction Coefficient", 0.9, "-", "Maps to Pacejka pDx1"),
        (None, "Longitudinal Friction Load Rating", 69.3, "kg", "Recommended value: M/4"),
        (None, "Longitudinal Friction Sensitivity", 0.0001, "1/N", None),
        (None, "Lateral Friction Coefficient", 1.5, "-", "Maps to Pacejka pDy1"),
        (None, "Lateral Friction Load Rating", 69.3, "kg", None),
        (None, "Lateral Friction Sensitivity", 0.0001, "1/N", None),
        (None, "Front Cornering Stiffness", 800, "N/deg", None),
        (None, "Rear Cornering Stiffness", 800, "N/deg", None),
        ("Engine", "Power Factor Multiplier", 1, "-", None),
        (None, "Thermal Efficiency", 0.95, "-", "EV inverter/motor (not used by fsae-6dof envelope)"),
        (None, "Fuel Lower Heating Value", 0, "J/kg", "Unused for EV"),
        ("Transmission", "Drive Type", "RWD", "-", None),
        (None, "Gear Shift Time", 0, "s", "Single-speed EV"),
        (None, "Primary Gear Efficiency", 1, "-", None),
        (None, "Final Gear Efficiency", 0.97, "-", None),
        (None, "Gearbox Efficiency", 1, "-", None),
        (None, "Primary Gear Reduction", 1, "-", "From motor to input shaft"),
        (None, "Final Gear Reduction", 4.8, "-", "Single-speed reduction (overridden by FastestLap Gear Ratio if set)"),
        (None, "1st Gear Ratio", 1, "-", "Single speed"),
        (None, "2nd Gear Ratio", None, "-", None),
        (None, "3rd Gear Ratio", None, "-", None),
        (None, "4th Gear Ratio", None, "-", None),
        (None, "5th Gear Ratio", None, "-", None),
        (None, "6th Gear Ratio", None, "-", None),
        (None, "7th Gear Ratio", None, "-", None),
        (None, "8th Gear Ratio", None, "-", None),
        (None, "9th Gear Ratio", None, "-", None),
        (None, "10th Gear Ratio", None, "-", None),
    ]


def ubco_2026_torque_rows() -> List[Tuple[float, float]]:
    # 80 kW / 240 N·m → base speed ≈ 3183 rpm. Torque falls as Pmax/ω above that.
    rows = [(1000, 240.0), (2000, 240.0), (3183, 240.0)]
    for rpm in (4000, 5000, 6000, 7000, 8000):
        omega = rpm * 2.0 * math.pi / 60.0
        rows.append((rpm, min(240.0, 80000.0 / omega)))
    return rows


def ubco_2026_fastestlap_rows() -> List[Tuple[Any, ...]]:
    return [
        ("Geometry", "Front Track", 1225, "mm", None),
        (None, "Rear Track", 1220, "mm", None),
        (None, "CG Height", 250, "mm", "Positive; stored as com.z = -h"),
        (None, "Front Axle Z", 47, "mm", "Chassis-frame axle height"),
        (None, "Rear Axle Z", 47, "mm", None),
        ("Suspension", "Front Wheel Rate", 130000, "N/m", None),
        (None, "Front Antiroll", 11260, "N/m", None),
        (None, "Front Damper", 1500, "N.s/m", None),
        (None, "Rear Wheel Rate", 143900, "N/m", None),
        (None, "Rear Antiroll", 11160, "N/m", None),
        (None, "Rear Damper", 1400, "N.s/m", None),
        (None, "Axle Inertia", 0.5, "kg.m2", "Per-wheel spin inertia"),
        ("Powertrain", "Peak Motor Torque", 240, "N.m", "Overrides Torque Curve peak if set"),
        (None, "Maximum Power", 80, "kW", "FSAE cap; overrides curve peak power if set"),
        (None, "Gear Ratio", 4.8, "-", "Motor to axle. Overrides primary×1st×final if set"),
        (None, "Regen Coefficient", 0.3, "-", "Negative throttle → regen * envelope"),
        (None, "Differential Stiffness", 80, "N.m.s/rad", None),
        ("Brakes", "Max Brake Torque", 1200, "N.m", "Per axle, scaled by brake bias"),
        (None, "Brake Bias", 0.53, "-", "1 = all front"),
        ("Kinematics", "Camber Static", -1.0, "deg", None),
        (None, "Camber Gain Roll", 0.3, "rad/rad", "Left/right opposite"),
        (None, "Toe Static", 0.0, "deg", None),
        (None, "Toe Gain Roll", 0.0, "rad/rad", None),
        ("Aero Maps", "dCl_dz", -8.0, "1/m", "Cl_eff = Cl*(1 + dCl_dz*z + dCl_dmu*mu)"),
        (None, "dCl_dmu", 2.0, "1/rad", None),
        (None, "dCd_dz", -2.0, "1/m", None),
        (None, "dCd_dmu", 0.5, "1/rad", None),
        ("Tire Thermal", "Tire Thermal Capacity", 900, "J/K", None),
        (None, "Tire Thermal Cooling", 15, "W/K", None),
        (None, "Tire T Ambient", 298.15, "K", None),
        (None, "Tire T Optimal", 353.15, "K", None),
        (None, "Tire Grip Sensitivity", 0.3, "-", None),
        ("Inertia", "Ixx", 30.0, "kg.m2", "Estimate"),
        (None, "Iyy", 90.0, "kg.m2", None),
        (None, "Izz", 110.0, "kg.m2", None),
        (None, "Ixz", 5.0, "kg.m2", None),
        ("Tyres", "Tire Radial Stiffness", 240000, "N/m", None),
        (None, "Tire Radial Damping", 1000, "N.s/m", None),
        (None, "Nominal Vertical Load", 680, "N", None),
    ]


def write_ubco_2026_xlsx(path: str | Path) -> Path:
    return write_openvehicle_workbook(
        path,
        ubco_2026_info_rows(),
        ubco_2026_torque_rows(),
        ubco_2026_fastestlap_rows(),
    )
