#!/usr/bin/env python3
"""OpenVEHICLE xlsx → fastest-lap XML importer tests."""

from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

_HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(_HERE.parent))

from fsae.openvehicle_xlsx import (  # noqa: E402
    derived_vehicle,
    read_openvehicle_xlsx,
    vehicle_to_xml,
    write_ubco_2026_xlsx,
    write_xml,
)


def _set_info_value(xlsx: Path, description: str, value) -> None:
    wb = load_workbook(xlsx)
    ws = wb["Info"]
    for row in ws.iter_rows(min_row=2):
        if row[1].value == description:
            row[2].value = value
            break
    else:
        raise AssertionError(f"{description} row missing")
    wb.save(xlsx)
    wb.close()


class TestOpenVehicleXlsx(unittest.TestCase):
    def test_parse_official_formula1_workbook(self) -> None:
        path = Path("/tmp/openvehicle/Formula 1.xlsx")
        if not path.is_file():
            self.skipTest("official OpenVEHICLE Formula 1.xlsx not present")
        data = read_openvehicle_xlsx(path)
        self.assertAlmostEqual(data["mass"], 650.0)
        self.assertAlmostEqual(data["wheelbase_mm"], 3000.0)
        self.assertAlmostEqual(data["cl_openvehicle"], -4.8)
        self.assertGreater(len(data["torque_curve"]), 10)
        self.assertAlmostEqual(data["torque_curve"][0][0], 1000.0)
        v = derived_vehicle(data)
        self.assertAlmostEqual(v["cl"], 4.8)
        self.assertAlmostEqual(v["gear_ratio"], 1.0 * 2.57 * 7.0)
        self.assertAlmostEqual(v["peak_torque"], 350.0)
        xml = vehicle_to_xml(v)
        self.assertIn('type="fsae-6dof"', xml)
        self.assertIn("<mass>650</mass>", xml)

    def test_ubco_workbook_roundtrip_mass_and_power(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            xlsx = Path(tmp) / "ubco.xlsx"
            xml_path = Path(tmp) / "ubco.xml"
            write_ubco_2026_xlsx(xlsx)
            data = read_openvehicle_xlsx(xlsx)
            self.assertAlmostEqual(data["mass"], 277.2)
            self.assertAlmostEqual(data["wheelbase_mm"], 1620.0)
            self.assertAlmostEqual(data["maximum_power_kw"], 80.0)
            self.assertAlmostEqual(data["peak_torque"], 240.0)
            v = derived_vehicle(data)
            self.assertAlmostEqual(v["mass"], 277.2)
            self.assertAlmostEqual(v["maximum_power_kw"], 80.0)
            self.assertAlmostEqual(v["da"], 0.48)
            self.assertAlmostEqual(v["camber_gain_roll"], 0.3)
            self.assertAlmostEqual(v["front_track_m"], 1.225)
            self.assertAlmostEqual(v["x_front"], (1.0 - 0.5127) * 1.62, places=6)
            self.assertAlmostEqual(v["x_rear"], -0.5127 * 1.62, places=6)
            write_xml(xlsx, xml_path)
            text = xml_path.read_text(encoding="utf-8")
            self.assertIn("<maximum-power units=\"kW\">80</maximum-power>", text)
            self.assertIn("<mass>277.2</mass>", text)
            self.assertIn("<camber_gain_roll>0.3</camber_gain_roll>", text)
            self.assertIn("<cl>3.77</cl>", text)
            self.assertIn("<cd>1.4</cd>", text)

    def test_changing_mass_in_xlsx_changes_xml(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            xlsx = Path(tmp) / "car.xlsx"
            xml_path = Path(tmp) / "car.xml"
            write_ubco_2026_xlsx(xlsx)
            _set_info_value(xlsx, "Total Mass", 301.5)
            write_xml(xlsx, xml_path)
            text = xml_path.read_text(encoding="utf-8")
            self.assertIn("<mass>301.5</mass>", text)
            self.assertNotIn("<mass>277.2</mass>", text)

    def test_openvehicle_cl_sign_is_flipped_to_downforce(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            xlsx = Path(tmp) / "ubco.xlsx"
            write_ubco_2026_xlsx(xlsx)
            data = read_openvehicle_xlsx(xlsx)
            self.assertLess(data["cl_openvehicle"], 0.0)
            v = derived_vehicle(data)
            self.assertGreater(v["cl"], 0.0)
            self.assertAlmostEqual(v["cl"], -data["cl_openvehicle"])

    def test_torque_curve_feeds_peak_when_fastestlap_omitted(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "openvehicle.xlsx"
            wb = Workbook()
            info = wb.active
            info.title = "Info"
            info.append(("Category", "Description", "Value", "Unit", "Comment"))
            info.append(("Inertia", "Total Mass", 280, "kg", None))
            info.append((None, "Front Mass Distribution", 50, "%", None))
            info.append(("Dimensions", "Wheelbase", 1600, "mm", None))
            info.append(("Aerodynamics", "Lift Coefficient CL", -2.0, "-", None))
            info.append((None, "Drag Coefficient CD", -1.0, "-", None))
            info.append((None, "Front Aero Distribution", 50, "%", None))
            info.append((None, "Frontal Area", 1.0, "m2", None))
            info.append(("Tyres", "Tyre Radius", 200, "mm", None))
            info.append(("Transmission", "Drive Type", "RWD", "-", None))
            info.append((None, "Primary Gear Reduction", 1.0, "-", None))
            info.append((None, "Final Gear Reduction", 5.0, "-", None))
            info.append((None, "1st Gear Ratio", 1.2, "-", None))
            torque = wb.create_sheet("Torque Curve")
            torque.append(("Engine Speed [rpm]", "Torque [Nm]"))
            torque.append((1000, 100))
            torque.append((4000, 180))
            wb.save(path)
            wb.close()

            v = derived_vehicle(read_openvehicle_xlsx(path))
            self.assertAlmostEqual(v["mass"], 280.0)
            self.assertAlmostEqual(v["peak_torque"], 180.0)
            self.assertAlmostEqual(v["gear_ratio"], 1.0 * 1.2 * 5.0)
            self.assertAlmostEqual(v["cl"], 2.0)
            self.assertAlmostEqual(v["cd"], 1.0)
            # 180 N·m at 4000 rpm = 75.398 kW
            self.assertAlmostEqual(v["maximum_power_kw"], 180.0 * 4000.0 * 2.0 * 3.141592653589793 / 60.0 / 1000.0, places=4)


if __name__ == "__main__":
    unittest.main()
