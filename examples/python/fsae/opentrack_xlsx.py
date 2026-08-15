"""Read an OpenTRACK-format workbook and emit a fastest-lap discrete circuit XML.

Workbook layout (same as OpenLAP / OpenTRACK Shape Data tmp.xlsx):

- Sheet ``Info``: Name, Country, City, Type, Configuration, Direction, Mirror
- Sheet ``Shape``: ``Type | Section Length | Corner Radius``
  Type is Straight / Left / Right. Radius 0 means a straight.
- Optional Elevation / Banking / Grip Factors / Sectors (ignored for 2D XML).

The Shape polyline is the racing line (OpenLAP does not store corridor width).
A constant half-width is applied so fastest-lap has left/right limits.
"""

from __future__ import annotations

import math
from dataclasses import dataclass
from pathlib import Path
from typing import Any, List, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

SHAPE_HEADER = ("Type", "Section Length", "Corner Radius")


@dataclass
class OpenTrackInfo:
    name: str = "Track"
    country: str = ""
    city: str = ""
    kind: str = "Temporary"
    configuration: str = "Closed"
    direction: str = "Forward"
    mirror: str = "Off"


@dataclass
class OpenTrackMesh:
    info: OpenTrackInfo
    s: List[float]
    x: List[float]
    y: List[float]
    yaw: List[float]
    kappa: List[float]
    nl: List[float]
    nr: List[float]

    @property
    def length(self) -> float:
        return float(self.s[-1]) if self.s else 0.0


def _cell_str(value: Any, default: str = "") -> str:
    if value is None:
        return default
    return str(value).strip().strip('"')


def read_opentrack_info(path: str | Path) -> OpenTrackInfo:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0] if "Info" not in wb.sheetnames else "Info"]
        kv = {}
        for row in ws.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            kv[_cell_str(row[0]).lower()] = _cell_str(row[1] if len(row) > 1 else None)
    finally:
        wb.close()
    return OpenTrackInfo(
        name=kv.get("name") or Path(path).stem,
        country=kv.get("country", ""),
        city=kv.get("city", ""),
        kind=kv.get("type", "Temporary"),
        configuration=kv.get("configuration", "Closed"),
        direction=kv.get("direction", "Forward"),
        mirror=kv.get("mirror", "Off"),
    )


def read_opentrack_shape(path: str | Path) -> List[Tuple[str, float, float]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        if "Shape" not in wb.sheetnames:
            raise ValueError(f"{path} has no 'Shape' sheet (OpenTRACK format).")
        rows: List[Tuple[str, float, float]] = []
        for i, row in enumerate(wb["Shape"].iter_rows(values_only=True), 1):
            if i == 1 or row is None or row[0] is None:
                continue
            kind = _cell_str(row[0]).title()
            length = float(row[1] or 0.0)
            radius = float(row[2] or 0.0)
            if length <= 0.0:
                continue
            rows.append((kind, length, radius))
    finally:
        wb.close()
    if not rows:
        raise ValueError(f"{path} Shape sheet has no segments.")
    return rows


def mesh_opentrack(
    path: str | Path,
    mesh_size: float = 1.0,
    half_width: float = 1.5,
) -> OpenTrackMesh:
    """Rebuild the OpenTRACK racing line at a uniform mesh (default 1 m).

    Closed tracks get OpenTRACK.m's linear start/finish XY correction so the
    map joins; yaw/kappa are then taken from the closed polyline.
    """
    info = read_opentrack_info(path)
    segments = read_opentrack_shape(path)
    closed = info.configuration.lower().startswith("closed")

    # Coarse curvature samples at segment midpoints (corners) and ends (straights).
    s_coarse: List[float] = []
    k_coarse: List[float] = []
    s_end = 0.0
    sign = -1.0 if info.mirror.lower() == "on" else 1.0
    for kind, length, radius in segments:
        k = 0.0
        if kind == "Left" and radius > 0.0:
            k = sign / radius
        elif kind == "Right" and radius > 0.0:
            k = -sign / radius
        if abs(k) < 1.0e-12:
            s_coarse.extend([s_end, s_end + length])
            k_coarse.extend([0.0, 0.0])
        else:
            s_coarse.append(s_end + 0.5 * length)
            k_coarse.append(k)
        s_end += length
    L = s_end
    if closed and (not s_coarse or s_coarse[-1] < L - 1.0e-9):
        s_coarse.append(L)
        k_coarse.append(k_coarse[0] if k_coarse else 0.0)
    s_coarse, k_coarse = _unique_samples(s_coarse, k_coarse)

    n = int(math.floor(L / mesh_size)) + 1
    s = [i * mesh_size for i in range(n)]
    if s[-1] < L - 1.0e-9:
        s.append(L)

    kappa = [_interp(s_coarse, k_coarse, si) for si in s]
    ds = [s[i + 1] - s[i] for i in range(len(s) - 1)]
    ds.append(ds[-1] if ds else mesh_size)

    heading = [0.0]
    for i in range(len(s) - 1):
        heading.append(heading[-1] + 0.5 * (kappa[i] + kappa[i + 1]) * ds[i])

    if closed and L > 0.0:
        wrap = heading[-1] - round(heading[-1] / (2.0 * math.pi)) * 2.0 * math.pi
        heading = [h - wrap * si / L for h, si in zip(heading, s)]
        heading = [h - heading[0] for h in heading]
        kappa = [0.0] + [(heading[i] - heading[i - 1]) / max(ds[i - 1], 1.0e-12) for i in range(1, len(s))]

    if info.direction.lower() == "backward":
        s = [L - si for si in reversed(s)]
        heading = [-h for h in reversed(heading)]
        kappa = [-k for k in reversed(kappa)]

    x = [0.0]
    y = [0.0]
    for i in range(len(s) - 1):
        step = s[i + 1] - s[i]
        x.append(x[-1] + step * math.cos(heading[i]))
        y.append(y[-1] + step * math.sin(heading[i]))

    if closed and L > 0.0 and len(x) > 1:
        dx_close = x[0] - x[-1]
        dy_close = y[0] - y[-1]
        x = [xi + si / L * dx_close for xi, si in zip(x, s)]
        y = [yi + si / L * dy_close for yi, si in zip(y, s)]
        x[-1], y[-1] = x[0], y[0]
        heading, kappa = _heading_kappa_from_xy(s, x, y, closed=True)

    nl = [half_width] * len(s)
    nr = [half_width] * len(s)
    return OpenTrackMesh(info=info, s=s, x=x, y=y, yaw=heading, kappa=kappa, nl=nl, nr=nr)


def _unique_samples(xp: Sequence[float], yp: Sequence[float]) -> Tuple[List[float], List[float]]:
    xs: List[float] = []
    ys: List[float] = []
    for x, y in zip(xp, yp):
        if xs and abs(x - xs[-1]) < 1.0e-12:
            ys[-1] = float(y)
            continue
        xs.append(float(x))
        ys.append(float(y))
    return xs, ys


def _wrap_pi(angle: float) -> float:
    while angle > math.pi:
        angle -= 2.0 * math.pi
    while angle < -math.pi:
        angle += 2.0 * math.pi
    return angle


def _heading_kappa_from_xy(
    s: Sequence[float], x: Sequence[float], y: Sequence[float], closed: bool
) -> Tuple[List[float], List[float]]:
    n = len(s)
    heading = [0.0] * n
    for i in range(n - 1):
        heading[i] = math.atan2(y[i + 1] - y[i], x[i + 1] - x[i])
    heading[-1] = heading[0] if closed and n > 1 else heading[-2] if n > 1 else 0.0
    for i in range(1, n):
        heading[i] = heading[i - 1] + _wrap_pi(heading[i] - heading[i - 1])
    kappa = [0.0] * n
    for i in range(1, n):
        kappa[i] = (heading[i] - heading[i - 1]) / max(s[i] - s[i - 1], 1.0e-12)
    if closed and n > 1:
        kappa[0] = kappa[-1]
    return heading, kappa


def _interp(xp: Sequence[float], yp: Sequence[float], x: float) -> float:
    if x <= xp[0]:
        return float(yp[0])
    if x >= xp[-1]:
        return float(yp[-1])
    for i in range(len(xp) - 1):
        if xp[i] <= x <= xp[i + 1]:
            t = 0.0 if xp[i + 1] == xp[i] else (x - xp[i]) / (xp[i + 1] - xp[i])
            return float(yp[i] + t * (yp[i + 1] - yp[i]))
    return float(yp[-1])


def _csv(values: Sequence[float], digits: int = 8) -> str:
    return ", ".join(f"{v:.{digits}g}" for v in values)


def mesh_to_discrete_xml(mesh: OpenTrackMesh) -> str:
    n = len(mesh.s)
    nx = [-math.sin(h) for h in mesh.yaw]
    ny = [math.cos(h) for h in mesh.yaw]
    lx = [xi + nli * nxi for xi, nli, nxi in zip(mesh.x, mesh.nl, nx)]
    ly = [yi + nli * nyi for yi, nli, nyi in zip(mesh.y, mesh.nl, ny)]
    rx = [xi - nri * nxi for xi, nri, nxi in zip(mesh.x, mesh.nr, nx)]
    ry = [yi - nri * nyi for yi, nri, nyi in zip(mesh.y, mesh.nr, ny)]
    dyaw = [0.0] + [
        (mesh.kappa[i] - mesh.kappa[i - 1]) / max(mesh.s[i] - mesh.s[i - 1], 1.0e-12)
        for i in range(1, n)
    ]
    zeros = [0.0] * n
    closed = "closed" if mesh.info.configuration.lower().startswith("closed") else "open"
    return f"""<circuit format="discrete" type="{closed}" dimensions="2">
    <header>
        <track_length units="m">{mesh.length:.8g}</track_length>
        <L2_error_left>0</L2_error_left>
        <L2_error_right>0</L2_error_right>
        <max_error_left>0</max_error_left>
        <max_error_right>0</max_error_right>
    </header>
    <optimization>
        <cost_curvature>10.000000</cost_curvature>
        <cost_track_limits_smoothness>0.000100</cost_track_limits_smoothness>
        <cost_track_limits_errors>0.000100</cost_track_limits_errors>
        <cost_centerline>0.000100</cost_centerline>
        <maximum_yaw_dot>4.000000</maximum_yaw_dot>
        <maximum_dyaw_dot>4.000000</maximum_dyaw_dot>
    </optimization>
    <GPS_parameters>
        <origin_longitude units="deg">0</origin_longitude>
        <origin_latitude units="deg">0</origin_latitude>
        <earth_radius units="m">6378388</earth_radius>
        <reference_latitude units="deg">0</reference_latitude>
    </GPS_parameters>
    <data number_of_points="{n}">
        <arclength units="m">{_csv(mesh.s)}</arclength>
        <centerline>
            <x units="m">{_csv(mesh.x)}</x>
            <y units="m">{_csv(mesh.y)}</y>
        </centerline>
        <left_boundary>
            <x units="m">{_csv(lx)}</x>
            <y units="m">{_csv(ly)}</y>
        </left_boundary>
        <right_boundary>
            <x units="m">{_csv(rx)}</x>
            <y units="m">{_csv(ry)}</y>
        </right_boundary>
        <left_measured_boundary>
            <x units="m">{_csv(lx)}</x>
            <y units="m">{_csv(ly)}</y>
        </left_measured_boundary>
        <right_measured_boundary>
            <x units="m">{_csv(rx)}</x>
            <y units="m">{_csv(ry)}</y>
        </right_measured_boundary>
        <yaw units="rad">{_csv(mesh.yaw)}</yaw>
        <yaw_dot units="rad">{_csv(mesh.kappa)}</yaw_dot>
        <nl units="m">{_csv(mesh.nl)}</nl>
        <nr units="m">{_csv(mesh.nr)}</nr>
        <dyaw_dot>{_csv(dyaw)}</dyaw_dot>
        <dnl>{_csv(zeros)}</dnl>
        <dnr>{_csv(zeros)}</dnr>
    </data>
</circuit>
"""


def write_map_xlsx(
    path: str | Path,
    info: OpenTrackInfo | dict,
    shape: Sequence[Tuple[str, float, float]],
) -> Path:
    """Write a Map workbook (Info + Shape). Used by Studio and tests."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    if isinstance(info, OpenTrackInfo):
        meta = info
    else:
        meta = OpenTrackInfo(
            name=str(info.get("name") or "Track"),
            country=str(info.get("country") or ""),
            city=str(info.get("city") or ""),
            kind=str(info.get("type") or info.get("kind") or "Temporary"),
            configuration=str(info.get("configuration") or "Closed"),
            direction=str(info.get("direction") or "Forward"),
            mirror=str(info.get("mirror") or "Off"),
        )
    wb = Workbook()
    ws = wb.active
    ws.title = "Info"
    for row in (
        ("Name", meta.name),
        ("Country", meta.country),
        ("City", meta.city),
        ("Type", meta.kind),
        ("Configuration", meta.configuration),
        ("Direction", meta.direction),
        ("Mirror", meta.mirror),
    ):
        ws.append(list(row))
    sh = wb.create_sheet("Shape")
    sh.append(list(SHAPE_HEADER))
    for kind, length, radius in shape:
        sh.append([str(kind), float(length), float(radius)])
    from fsae.xlsx_kit import stamp_format

    stamp_format(wb, "map")
    wb.save(path)
    return path


def write_discrete_xml(xlsx_path: str | Path, xml_path: str | Path, **kwargs) -> Path:
    xml_path = Path(xml_path)
    xml_path.parent.mkdir(parents=True, exist_ok=True)
    xml_path.write_text(mesh_to_discrete_xml(mesh_opentrack(xlsx_path, **kwargs)), encoding="utf-8")
    return xml_path


def write_fsae_skidpad_xlsx(path: str | Path) -> Path:
    """FSAE figure-8 skidpad: two 9.125 m radius circles (OpenTRACK Shape layout)."""
    return write_map_xlsx(
        path,
        OpenTrackInfo(name="FSAE Skidpad", country="Greece", city="Athens"),
        (
            ("Left", 1, 9.125),
            ("Left", 55, 9.125),
            ("Left", 1, 9.125),
            ("Right", 1, 9.125),
            ("Right", 55, 9.125),
            ("Right", 1, 9.125),
        ),
    )
